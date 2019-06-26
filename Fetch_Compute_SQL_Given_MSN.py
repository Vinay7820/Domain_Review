import pandas as pd
import argparse, ast, boto3, json, os, pyodbc, subprocess, sys, shutil
import openpyxl
from openpyxl import load_workbook


lines = []
no_compute = []
row = 1

def search_file(msn, path):
    global row
    filename = path+msn
    print("File = ", filename)
    for root, dirs, files in os.walk(path):
        if msn in files:
            print("Found File")
            row = row+1
            with open(filename) as f:
                s = f.read()
                print(s)
                writetoexcel(s, row)
        else:
            row = row+1
            writetoexcel("NULL", row)


def writetoexcel(lines, row):
    workbook_name = 'Fetch_MSN_Month.xlsx'
    wb = load_workbook(workbook_name)
    ws1 = wb['Sheet1']
    row_no = row
    column_no = 2
    ws1.cell(row=row_no, column=column_no).value = lines
    wb.save(filename=workbook_name)

if __name__ == '__main__':
    search_path = 'C:/Users/vinay.m/PycharmProjects/Learnings/Common-Queries/consumer_metric_queries/monthly/compute_sql/'
    fields = ['MSN']
    df = pd.read_csv("Fetch_MSN_Month.csv", skipinitialspace=True, usecols=fields)
    for i in df.MSN:
        print("\n")
        print("MSN = ", i)
        msn = "compute_"+i+".sql"
        print("Searching the compute SQL = ",i)
        find_file = search_file(msn,search_path)


