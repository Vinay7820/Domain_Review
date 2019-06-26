# encoding=utf8
import re
import time
import smtplib
from jira import JIRA
from datetime import datetime
import csv
import openpyxl
import boto3
import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

# reload(sys)
# sys.setdefaultencoding('utf8')

wb = openpyxl.Workbook()
sheet = wb.active
wb.create_sheet(index=1, title="jira_pull")
wb.save("Jira_pull.xlsx")
column_name = ['key', 'summary', 'created', 'assignee', 'updated', 'reporter', 'priority', 'status',
                      'resolution', 'timespent', 'subtasks', 'issuetype', 'labels', 'datacategory', 'Component/s',
                      'Description']
workbook_name = 'Jira_pull.xlsx'
wb = load_workbook(workbook_name)
ws1 = wb['jira_pull']
#max = ws1.max_row
start_column = 1
for header in column_name:
    ws1.cell(row=1, column=start_column).value = header
    start_column += 1
wb.save(filename=workbook_name)

Year = datetime.now().strftime('%Y')
Month = datetime.now().strftime('%m')
Day = datetime.now().strftime('%d')
data_category = ""
row = 2


def writetoexcel(row_value, components, issue):
    column_value = 1
    workbook_name = 'Jira_pull.xlsx'
    wb = load_workbook(workbook_name)
    ws1 = wb['jira_pull']
    ws1.cell(row=row_value, column=column_value).value = str(issue.key)
    ws1.cell(row=row_value, column=column_value+1).value = str(issue.fields.summary)
    ws1.cell(row=row_value, column=column_value+2).value = str(issue.fields.created)
    ws1.cell(row=row_value, column=column_value+3).value = str(issue.fields.assignee)
    ws1.cell(row=row_value, column=column_value+4).value = str(issue.fields.updated)
    ws1.cell(row=row_value, column=column_value+5).value = str(issue.fields.reporter)
    ws1.cell(row=row_value, column=column_value+6).value = str(issue.fields.priority)
    ws1.cell(row=row_value, column=column_value+7).value = str(issue.fields.status)
    ws1.cell(row=row_value, column=column_value+8).value = str(issue.fields.resolution)
    ws1.cell(row=row_value, column=column_value+9).value = str(issue.fields.timespent)
    ws1.cell(row=row_value, column=column_value+10).value = str(issue.fields.subtasks)
    ws1.cell(row=row_value, column=column_value+11).value = str(issue.fields.issuetype)
    ws1.cell(row=row_value, column=column_value+12).value = str(issue.fields.labels)
    ws1.cell(row=row_value, column=column_value+13).value = str(issue.fields.customfield_13004)
    ws1.cell(row=row_value, column=column_value+14).value = str(components)
    ws1.cell(row=row_value, column=column_value+15).value = str(issue.fields.description)
    wb.save(filename=workbook_name)


try:
    print("Trying to Login into JIRA")
    Jira = JIRA(basic_auth=('jira_monitoring', 'jira_monitoring'), options={'server': 'https://jira.move.com'})
    print("Logged into JIRA succesfully")
except Exception as e:
    print("Not able to login into JIRA. Please check the connection and try again!")
issues_in_project = Jira.search_issues('project=DE AND created >= 2019-03-01', maxResults=None)
for issue in issues_in_project:
    # print(issue.fields.customfield_13004)
    fieldnames = ['key', 'summary', 'created', 'assignee', 'updated', 'reporter', 'priority', 'status',
                  'resolution', 'timespent', 'subtasks', 'issuetype', 'labels', 'datacategory', 'Component/s',
                  'Description']
    #writer = csv.DictWriter(csvFile, fieldnames=fieldnames)
    if str(issue.fields.components) != '[]':
        component = str(issue.fields.components)
        components = component.split("', id=", 1)[0].split("<JIRA Component: name='", 1)[1]
    else:
        components = 'None'


    writetoexcel(row, components, issue)
    row += 1
print("Excel File Generated")

pd.read_excel('Jira_pull.xlsx', sheet_name='jira_pull').to_csv('Jira_dump.csv', index=False)

print("Connecting to s3")
s3 = boto3.resource('s3')
s3.meta.client.upload_file('./Jira_dump.csv', 'move-dataeng-temp-dev', 'Jira/Jira_dump.csv')
#s3.meta.client.upload_file('./Jira_dump.csv', 'move-dataeng-temp-prod', 'Jira_history/'+str(Year)+'/'+str(Month)+'/'+str(Day)+'/'+'Jira_history.csv')
print("csv file has been Copied to s3")